# -*- coding: utf-8 -*-
"""Exportación del listado de instancias a Excel (.xlsx) y CSV."""
from __future__ import annotations

import csv
import io

# Columnas: (clave, titulo, ancho)
COLUMNAS = [
    ('cliente', 'Cliente', 16),
    ('tipo', 'Sistema', 12),
    ('empresa', 'Empresa', 30),
    ('ruc', 'RUC', 16),
    ('ruc_proveedor', 'RUC facturador', 18),
    ('url', 'URL', 34),
    ('url_estado', 'URL responde', 14),
    ('dominio_credenciales', 'DOMINIO_GENERAL (credenciales)', 30),
    ('dominio_desactualizado', 'Dominio desactualizado', 20),
    ('servicio', 'Servicio systemd', 20),
    ('servicio_estado', 'Estado servicio', 15),
    ('servicio_uptime', 'Uptime', 12),
    ('cpu_pct', 'CPU %', 9),
    ('ram_legible', 'RAM', 11),
    ('ram_pct', 'RAM % servidor', 15),
    ('servicio_desde', 'Activo desde', 18),
    ('servicio_creado', 'Servicio creado', 17),
    ('fecha_instalacion', 'Instalado el', 17),
    ('ruta', 'Ruta de instalación', 38),
    ('servidor_web', 'Servidor web', 14),
    ('servidor_web_activo', 'Servidor web activo', 18),
    ('apache_archivo', 'Archivo del sitio', 42),
    ('apache_habilitado', 'Sitio habilitado', 16),
    ('ssl_estado', 'SSL', 12),
    ('ssl_hasta', 'SSL vence', 13),
    ('ssl_dias', 'SSL días', 10),
    ('ssl_emisor', 'SSL emisor', 22),
    ('db_nombre', 'Base de datos', 20),
    ('db_host', 'Host BD', 16),
    ('db_estado', 'Estado BD', 12),
    ('db_tamano', 'Tamaño BD', 13),
    ('media_tamano', 'Tamaño media', 13),
    ('db_pct_disco', 'BD % del disco', 14),
    ('media_pct_disco', 'Media % del disco', 17),
    ('logs_tamano', 'Tamaño logs', 13),
    ('ocupa_legible', 'Ocupa total', 13),
    ('ocupa_pct_disco', 'Ocupa % del disco', 17),
    ('logs_archivos', 'Archivos de log', 15),
    ('dias_sin_uso', 'Días sin uso', 13),
    ('api_cedula', 'API cédula', 12),
    ('auditoria_fecha', 'Última auditoría', 17),
    ('auditoria_usuario', 'Usuario auditoría', 18),
    ('auditoria_accion', 'Acción auditoría', 16),
    ('auditoria_tabla', 'Tabla auditada', 26),
    ('ultima_sesion', 'Última sesión', 19),
    ('ultima_sesion_usuario', 'Usuario última sesión', 20),
    ('facturas_total', 'Facturas totales', 15),
    ('facturas_mes', 'Facturas mes actual', 18),
    ('facturas_mes_anterior', 'Facturas mes anterior', 19),
    ('facturas_ultimo_mes', 'Último mes facturado', 19),
    ('facturas_meses_sin', 'Meses sin facturar', 17),
    ('primera_venta', 'Primera venta', 14),
    ('ultima_venta', 'Última venta', 14),
    ('ventas_total', 'Total ventas', 13),
]


def fila(inst):
    """Aplana una instancia a las columnas de exportación."""
    r = inst.get('resumen') or {}
    db = inst.get('db') or {}
    return {
        'cliente': inst.get('cliente'),
        'tipo': inst.get('tipo'),
        'empresa': r.get('empresa'),
        'ruc': r.get('ruc'),
        'ruc_proveedor': (r.get('ruc_proveedor') if r.get('ruc_proveedor_disponible')
                          else 'sin el campo'),
        'url': inst.get('url'),
        'url_estado': ('SI (HTTP %s)' % r.get('url_codigo')) if r.get('url_responde')
                      else ('NO' if r.get('url_responde') is False else '-'),
        'servicio': inst.get('servicio'),
        'servicio_estado': r.get('servicio_estado'),
        'servicio_uptime': r.get('servicio_uptime'),
        'cpu_pct': r.get('cpu_pct'),
        'ram_legible': r.get('ram_legible'),
        'ram_pct': r.get('ram_pct'),
        'servicio_desde': r.get('servicio_desde'),
        'servicio_creado': r.get('servicio_creado'),
        'fecha_instalacion': r.get('fecha_instalacion'),
        'ruta': inst.get('ruta'),
        'servidor_web': r.get('servidor_web'),
        'servidor_web_activo': ('SI' if r.get('servidor_web_activo') else
                                ('NO' if r.get('servidor_web_activo') is False else '-')),
        'apache_archivo': r.get('apache_archivo'),
        'apache_habilitado': ('SI' if r.get('apache_habilitado')
                              else ('NO' if r.get('apache_archivo') else '-')),
        'ssl_estado': r.get('ssl_estado'),
        'ssl_hasta': r.get('ssl_hasta'),
        'ssl_dias': r.get('ssl_dias'),
        'ssl_emisor': r.get('ssl_emisor'),
        'db_nombre': db.get('dbname'),
        'db_host': db.get('host'),
        'db_estado': ('-' if db.get('desactivado') else ('activa' if db.get('ok') else 'CAIDA')),
        'db_tamano': r.get('db_tamano'),
        'media_tamano': r.get('media_tamano'),
        'db_pct_disco': r.get('db_pct_disco'),
        'media_pct_disco': r.get('media_pct_disco'),
        'ocupa_legible': r.get('ocupa_legible'),
        'ocupa_pct_disco': r.get('ocupa_pct_disco'),
        'logs_tamano': r.get('logs_tamano'),
        'logs_archivos': r.get('logs_archivos'),
        'dias_sin_uso': r.get('dias_sin_uso'),
        'api_cedula': ('activa' if r.get('api_cedula') else
                       ('inactiva' if r.get('api_cedula') is False else '-')),
        'auditoria_fecha': r.get('auditoria_fecha'),
        'auditoria_usuario': r.get('auditoria_usuario'),
        'auditoria_accion': {'A': 'Adición', 'M': 'Modificación', 'E': 'Eliminación'}
                            .get(r.get('auditoria_accion'), r.get('auditoria_accion')),
        'auditoria_tabla': r.get('auditoria_tabla'),
        'ultima_sesion': r.get('ultima_sesion'),
        'ultima_sesion_usuario': r.get('ultima_sesion_usuario'),
        'dominio_credenciales': r.get('dominio_credenciales'),
        'dominio_desactualizado': 'SI' if r.get('dominio_desactualizado') else 'no',
        'facturas_total': r.get('facturas_total'),
        'facturas_mes': r.get('facturas_mes'),
        'facturas_mes_anterior': r.get('facturas_mes_anterior'),
        'facturas_ultimo_mes': r.get('facturas_ultimo_mes'),
        'facturas_meses_sin': r.get('facturas_meses_sin'),
        'primera_venta': r.get('primera_venta'),
        'ultima_venta': r.get('ultima_venta'),
        'ventas_total': r.get('ventas_total'),
    }


def a_csv(instancias):
    """CSV con separador ';' y BOM, listo para abrirse en Excel."""
    buffer = io.StringIO()
    escritor = csv.writer(buffer, delimiter=';')
    escritor.writerow([titulo for _, titulo, _ in COLUMNAS])
    for inst in instancias:
        datos = fila(inst)
        escritor.writerow([datos.get(clave, '') for clave, _, _ in COLUMNAS])
    return '\ufeff' + buffer.getvalue()


def a_xlsx(instancias, resumen=None, titulo='Instancias'):
    """Genera el .xlsx en memoria. Devuelve (bytes, None) o (None, error)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as ex:
        return None, ('openpyxl no está instalado (%s). Ejecuta: '
                      'venv/bin/pip install openpyxl' % ex)

    libro = Workbook()
    hoja = libro.active
    hoja.title = 'Instancias'

    relleno_cabecera = PatternFill('solid', fgColor='12325F')
    fuente_cabecera = Font(color='FFFFFF', bold=True, size=10)
    verde = PatternFill('solid', fgColor='E2F6ED')
    rojo = PatternFill('solid', fgColor='FDECEB')
    ambar = PatternFill('solid', fgColor='FDF3E0')

    hoja.append([t for _, t, _ in COLUMNAS])
    for celda in hoja[1]:
        celda.fill = relleno_cabecera
        celda.font = fuente_cabecera
        celda.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
    hoja.row_dimensions[1].height = 28

    indices = {clave: i + 1 for i, (clave, _, _) in enumerate(COLUMNAS)}
    for inst in instancias:
        datos = fila(inst)
        hoja.append([datos.get(clave) for clave, _, _ in COLUMNAS])
        n = hoja.max_row

        estado_servicio = (datos.get('servicio_estado') or '')
        hoja.cell(row=n, column=indices['servicio_estado']).fill = (
            verde if estado_servicio in ('active', 'activating') else rojo)

        hoja.cell(row=n, column=indices['url_estado']).fill = (
            verde if str(datos.get('url_estado', '')).startswith('SI') else
            (rojo if datos.get('url_estado') == 'NO' else ambar))

        ssl_estado = datos.get('ssl_estado')
        hoja.cell(row=n, column=indices['ssl_estado']).fill = (
            verde if ssl_estado == 'vigente' else
            (ambar if ssl_estado == 'por-vencer' else rojo))

        hoja.cell(row=n, column=indices['apache_habilitado']).fill = (
            verde if datos.get('apache_habilitado') == 'SI' else
            (rojo if datos.get('apache_habilitado') == 'NO' else ambar))

        meses = datos.get('facturas_meses_sin')
        celda_fact = hoja.cell(row=n, column=indices['facturas_mes'])
        if datos.get('facturas_mes'):
            celda_fact.fill = verde
        elif meses is None:
            pass
        elif meses <= 1:
            celda_fact.fill = ambar
        else:
            celda_fact.fill = rojo

        if datos.get('dominio_desactualizado') == 'SI':
            hoja.cell(row=n, column=indices['dominio_desactualizado']).fill = ambar

        estado_db = datos.get('db_estado')
        if estado_db and estado_db != '-':
            hoja.cell(row=n, column=indices['db_estado']).fill = (
                verde if estado_db == 'activa' else rojo)

    for i, (_, _, ancho) in enumerate(COLUMNAS, start=1):
        hoja.column_dimensions[get_column_letter(i)].width = ancho
    hoja.freeze_panes = 'C2'
    hoja.auto_filter.ref = 'A1:%s%s' % (get_column_letter(len(COLUMNAS)), hoja.max_row)

    if resumen:
        hoja2 = libro.create_sheet('Resumen')
        hoja2.append(['Indicador', 'Valor'])
        for celda in hoja2[1]:
            celda.fill = relleno_cabecera
            celda.font = fuente_cabecera
        etiquetas = [
            ('Instancias totales', resumen.get('total')),
            ('Servicios activos', resumen.get('servicios_activos')),
            ('Servicios inactivos', resumen.get('servicios_inactivos')),
            ('Sitios Apache habilitados', resumen.get('sitios_habilitados')),
            ('Certificados vigentes', resumen.get('ssl_vigentes')),
            ('Certificados vencidos o por vencer', resumen.get('ssl_alerta')),
            ('URLs que responden', resumen.get('urls_ok')),
            ('Bases accesibles', resumen.get('db_activas')),
            ('Tamaño total de bases', resumen.get('db_tamano')),
            ('Tamaño total de media', resumen.get('media_tamano')),
            ('Tamaño total de logs', resumen.get('logs_tamano')),
            ('RAM usada por las instancias', resumen.get('ram_tamano')),
            ('CPU usada por las instancias (% de un núcleo)', resumen.get('cpu_pct')),
            ('Ocupación total (BD + media + logs)', resumen.get('ocupa_tamano')),
        ]
        for tipo, cantidad in (resumen.get('por_tipo') or {}).items():
            etiquetas.append(('Instancias de %s' % tipo, cantidad))
        for etiqueta, valor in etiquetas:
            hoja2.append([etiqueta, valor])
        hoja2.column_dimensions['A'].width = 38
        hoja2.column_dimensions['B'].width = 18

    flujo = io.BytesIO()
    libro.save(flujo)
    return flujo.getvalue(), None
